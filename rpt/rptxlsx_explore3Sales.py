import openpyxl
import fxlpalmtree
from pprint import pprint
from copy import copy

# wb = openpyxl.load_workbook('Report Inventory Recipe Forecast.xlsx', read_only=False)
wb = openpyxl.load_workbook('nested-sales.xlsx')
tpl = wb.active

# wb2 = openpyxl.Workbook()
# ws = wb2.active 
wb2 = wb

ws = wb2.create_sheet()

# print( dir(cell.column) )
# for (row, col), source_cell  in tpl._cells.items():
#     print('!', row, col)
first = True
for A, cd in tpl.column_dimensions.items():
    if first:
        first = False
        continue
    B = chr(ord(A)-1)
    ws.column_dimensions[B].width = cd.width
ws.sheet_format = copy(tpl.sheet_format)

palm = fxlpalmtree.palmTree()
palm.arrangeSeed(tpl)
# pprint( palm.showMap() )

sheet_name = tpl.title
wb.remove(tpl)
tpl = None
ws.title = sheet_name

from Items import getItemByName, getItemById

class myItems(fxlpalmtree.fxl):
    _name = 'Item'

    #def onCreate(self):
    def init( self ): #its call before any worksheet parsed.
        # from Items import dats as items
        from Sellhistory import dats as solds
        for sold in solds: # each row
            item_id = sold['Name']
            item = getItemById(item_id)
            sold.update(item)
        # self.data = dats
        self.datas = solds
        
myItems()        
        
class myBom(fxlpalmtree.fxl):
    _name = 'BoM'

    #def onCreate(self):
    def init( self ): #its call before any worksheet parsed.
        item = fxlpalmtree.BANDS['Item'].data
        data =[]
        for b in item.get('BoM',[]):
            name,qty = b
            row = getItemByName(name)
            row['RecipeQty'] = float(qty)
            row['SellQuantity'] = item['SellQuantity']
            row['SellUoM'] = item['Measure']
            data.append(row)
            
        # print('BoM.data=', data)
        self.datas = data
    



myBom()
palm.harvestFarm(ws)

wb2.save('nested-sales-out.xlsx')

# for i, cd in ws1.column_dimensions.items():
#     print(i, cd.width, cd.index)
#     # ws2.column_dimensions[k].width = cd.width
# print('***', dir(cd))
# # from openpyxl.utils import get_column_letter

# column_widths = []
# for row in ws1.rows:
#     for i, cell in enumerate(row):
#         if len(column_widths) > i:
#             if len(cell) > column_widths[i]:
#                 column_widths[i] = len(cell)
#         else:
#             column_widths += [len(cell)]
#     break

# ws = your current worksheet
# dims = {}
# for row in ws1.rows:
#     for cell in row:
#         if cell.value:
#             # dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value)))) 
#             dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))   
#             # dims[cell.column_letter] = max((dims.get(cell.column_letter, 0))   
# for col, value in dims.items():
#     # ws1.column_dimensions[col].width = value
#     print(col,'$',ws1.column_dimensions[col].width )

# for i, column_width in enumerate(column_widths):
#     # worksheet.column_dimensions[get_column_letter(i+1)].width = column_width
#     print( column_width )
    
# for row in ws1.iter_rows():
# for row in tpl.rows:
# #     print(row, dir(row) )
# #     # break
#     for i,cell in enumerate(row):
#         print(cell, dir(cell) )
# #         # print( cell.column_letter )
# #         # col = 'A'+str(cell.column)
# #         col = chr(65+i)
# #         print( col, ws1.column_dimensions[col].width, ws1.column_dimensions[col].customWidth )
# #         # print( dir(cell.column) )
# #         # print( dir(cell) )
# #         # print( (cell.column) )
#         break
#     break