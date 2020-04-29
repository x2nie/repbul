import openpyxl as xls

# https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/worksheet/copier.html

wbIn = xls.load_workbook('Inventory Report with  Sub-Category.xlsx')
wbOut = xls.Workbook()

tpl = wbIn.active
ws2 = wbOut.active
# print( dir(tpl) )
print( dir(tpl.sheet_format) )
print( 'COLS',[c for c in tpl.columns] )

for i, cd in tpl.column_dimensions.items():
    print(i, cd.width, cd.index)
    ws2.column_dimensions[i].width = cd.width
    
def iter_rows(ws,n):  #produce the list of items in the particular row
        for row in ws.iter_rows(n):
            yield [cell.value for cell in row]

for row in tpl.iter_rows():
    ws2.append([cell.value for cell in row])
    
wbOut.save(filename = 'dest_filename.xlsx')
# print('***', dir(cd))
# from openpyxl.utils import get_column_letter

# column_widths = []
# for row in ws1.rows:
#     for i, cell in enumerate(row):
#         if len(column_widths) > i:
#             if len(cell) > column_widths[i]:
#                 column_widths[i] = len(cell)
#         else:
#             column_widths += [len(cell)]
#     break

# ws = your current worksheet
# dims = {}
# for row in ws1.rows:
#     for cell in row:
#         if cell.value:
#             # dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value)))) 
#             dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))   
#             # dims[cell.column_letter] = max((dims.get(cell.column_letter, 0))   
# for col, value in dims.items():
#     # ws1.column_dimensions[col].width = value
#     print(col,'$',ws1.column_dimensions[col].width )

# for i, column_width in enumerate(column_widths):
#     # worksheet.column_dimensions[get_column_letter(i+1)].width = column_width
#     print( column_width )
    
# for row in ws1.iter_rows():
# for row in ws1.rows:
#     print(row, dir(row) )
#     for i,cell in enumerate(row):
#         # print(cell, dir(cell) )
#         # print( cell.column_letter )
#         # col = 'A'+str(cell.column)
#         col = chr(65+i)
#         print( col, ws1.column_dimensions[col].width, ws1.column_dimensions[col].customWidth )
#         # print( dir(cell.column) )
#         # print( dir(cell) )
#         # print( (cell.column) )
#         # break
#     break