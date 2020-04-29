import openpyxl

tpl = openpyxl.load_workbook('Inventory Report with  Sub-Category.xlsx')
# wb2 = openpyxl.load_workbook('file2.xlsx')

ws1 = tpl.active
# ws2 = wb2.active
# print( dir(cell.column) )

from openpyxl.utils import get_column_letter

# column_widths = []
# for row in ws1.rows:
#     for i, cell in enumerate(row):
#         if len(column_widths) > i:
#             if len(cell) > column_widths[i]:
#                 column_widths[i] = len(cell)
#         else:
#             column_widths += [len(cell)]
#     break

# ws = your current worksheet
dims = {}
for row in ws1.rows:
    for cell in row:
        if cell.value:
            # dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value)))) 
            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))   
for col, value in dims.items():
    ws1.column_dimensions[col].width = value

for i, column_width in enumerate(column_widths):
    # worksheet.column_dimensions[get_column_letter(i+1)].width = column_width
    print( column_width )
    
# for row in ws1.iter_rows():
for row in ws1.rows:
    for cell in row:
        print( cell.column_letter )
        col = 'A'+str(cell.column)
        print( col, ws1.column_dimensions[col].width )
        # print( dir(cell.column) )
        # print( dir(cell) )
        # print( (cell.column) )
        # break
    break