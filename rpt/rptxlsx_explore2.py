import openpyxl
import fxlpalmtree
from pprint import pprint

# wb = openpyxl.load_workbook('Report Inventory Recipe Forecast.xlsx', read_only=False)
wb = openpyxl.load_workbook('flat.xlsx')
wb2 = openpyxl.Workbook()

tpl = wb.active
ws = wb2.active 
# print( dir(cell.column) )
# for (row, col), source_cell  in tpl._cells.items():
#     print('!', row, col)

palm = fxlpalmtree.palmTree()
palm.arrangeSeed(tpl)
pprint( palm.showMap() )

class myItems(fxlpalmtree.fxl):
    _name = 'Item'

    #def onCreate(self):
    def init( self ): #its call before any worksheet parsed.
        from Items import dats
        self.data = dats
    
    def hasData( self ):
        "Needed for prevent endless loop"
        return self.counter < len(self.data)
    
    def reset( self ):
        self.counter = 0
        self.init()
        
    def bandTitle( self ):
        #canReturn= ['How Much','Where','When']
        return []
    
    def bandType( self ):
        #canReturn= ['Number','String','DateTime']
        return []
    
    def bandData( self ):
        if self.hasData():
            dat = self.data[self.counter]
            self.counter += 1
            return dat
            ##t=time.time()
            # t = datetime.fromtimestamp( time.time() )
            # t = t.strftime( '%Y-%m-%dT%H:%M:%S' )
            # return {'No.':self.counter, 'str':'Number' + str( self.counter ), 'today':t}
        else:
            return {}

myItems()
palm.harvestFarm(ws)

wb2.save('flat-out.xlsx')

# for i, cd in ws1.column_dimensions.items():
#     print(i, cd.width, cd.index)
#     # ws2.column_dimensions[k].width = cd.width
# print('***', dir(cd))
# # from openpyxl.utils import get_column_letter

# column_widths = []
# for row in ws1.rows:
#     for i, cell in enumerate(row):
#         if len(column_widths) > i:
#             if len(cell) > column_widths[i]:
#                 column_widths[i] = len(cell)
#         else:
#             column_widths += [len(cell)]
#     break

# ws = your current worksheet
# dims = {}
# for row in ws1.rows:
#     for cell in row:
#         if cell.value:
#             # dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value)))) 
#             dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))   
#             # dims[cell.column_letter] = max((dims.get(cell.column_letter, 0))   
# for col, value in dims.items():
#     # ws1.column_dimensions[col].width = value
#     print(col,'$',ws1.column_dimensions[col].width )

# for i, column_width in enumerate(column_widths):
#     # worksheet.column_dimensions[get_column_letter(i+1)].width = column_width
#     print( column_width )
    
# for row in ws1.iter_rows():
# for row in tpl.rows:
# #     print(row, dir(row) )
# #     # break
#     for i,cell in enumerate(row):
#         print(cell, dir(cell) )
# #         # print( cell.column_letter )
# #         # col = 'A'+str(cell.column)
# #         col = chr(65+i)
# #         print( col, ws1.column_dimensions[col].width, ws1.column_dimensions[col].customWidth )
# #         # print( dir(cell.column) )
# #         # print( dir(cell) )
# #         # print( (cell.column) )
#         break
#     break